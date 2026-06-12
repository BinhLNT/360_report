# -*- coding: utf-8 -*-
"""
competency_exporter.py
======================
Sinh "FILE THỨ 4" ở ĐỊNH DẠNG WIDE, BÁM ĐÚNG bố cục file gốc
"360 data raw(Sheet1).csv" (đã đo chính xác bằng _analyze_sheet1):

  c1-6   Định danh: Mã NV | Họ và tên | Chức danh | Ban/Chuỗi/Khối | Cấp bậc | Trạng thái
  c7     KẾT QUẢ ĐÁNH GIÁ
  c8-55  TỔNG ĐIỂM THEO TIÊU CHÍ  — 24 hành vi × 2 cột (Điểm + Diễn giải)
  c56-79 CẤP TRÊN                 — 24 hành vi × 1 cột (Điểm)
  c80-103 ĐỒNG NGHIỆP/ ĐỐI TÁC    — 24 × 1
  c104-127 CẤP DƯỚI               — 24 × 1
  Khuyến nghị
  -> sau đó NỐI THÊM: [16 cột AI] + [4 cột rà soát]  (cột AI nằm cuối).

Header 3 hàng: banner khối (gộp) → tên nhóm năng lực (gộp) → mô tả hành vi.
Điểm từng hành vi lấy từ structured["behaviors"], khớp 24 hành vi chuẩn theo
văn bản (ưu tiên) rồi theo thứ tự trong nhóm.
"""

import csv
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import config
import score_calculator as calc


def _fmt(value, decimals=config.DISPLAY_DECIMALS):
    if value is None:
        return ""
    return round(float(value), decimals)


def _norm(text):
    """Chuẩn hoá text để so khớp hành vi (bỏ dấu, gộp khoảng trắng)."""
    return " ".join(calc.strip_accents(text).split())


# ---------------------------------------------------------------------------
# Khớp điểm hành vi của 1 nhân viên vào 24 hành vi chuẩn
# ---------------------------------------------------------------------------
def _match_behavior_scores(structured):
    """
    Trả về list 24 dict điểm {others, cap_tren, dong_cap, cap_duoi} đúng thứ tự
    config.BEHAVIORS.

    Khớp 2 LƯỢT, KHÔNG để 1 hành vi engine bị gán cho 2 ô (no double-assignment):
      - Lượt 1: khớp theo VĂN BẢN chuẩn hoá — ưu tiên toàn cục, chỉ nhận hành vi
        chưa dùng (xử lý cả trường hợp nhiều hành vi cùng văn bản bằng hàng đợi).
      - Lượt 2: các ô còn trống lấy theo THỨ TỰ trong cùng subcomp_key, cũng chỉ
        nhận hành vi chưa dùng.
    """
    behaviors = structured.get("behaviors", [])
    by_text, by_subcomp = {}, {}
    for b in behaviors:
        by_text.setdefault(_norm(b["behavior"]), []).append(b)   # hàng đợi (chịu trùng văn bản)
        by_subcomp.setdefault(b["subcomp_key"], []).append(b)
    used = set()
    matched = [None] * len(config.BEHAVIORS)

    # Lượt 1: khớp văn bản (ưu tiên toàn cục, bỏ hành vi đã dùng).
    for i, (_sub_key, canonical_text) in enumerate(config.BEHAVIORS):
        for cand in by_text.get(_norm(canonical_text), []):
            if id(cand) not in used:
                used.add(id(cand))
                matched[i] = cand
                break

    # Lượt 2: fallback theo thứ tự trong cùng nhóm, cho các ô còn trống.
    for i, (sub_key, _canonical_text) in enumerate(config.BEHAVIORS):
        if matched[i] is not None:
            continue
        for cand in by_subcomp.get(sub_key, []):
            if id(cand) not in used:
                used.add(id(cand))
                matched[i] = cand
                break

    return [
        {
            **{k: (m.get(k) if m else None) for k in ("others", "cap_tren", "dong_cap", "cap_duoi")},
            "comments": (m.get("comments") if m else None) or [],
        }
        for m in matched
    ]


def _format_behavior_comments(comments):
    """Định dạng ý kiến theo từng mục tiêu cho cột 'Diễn giải' (mỗi người 1 dòng,
    ẩn danh theo nhóm quan hệ, sắp xếp Cấp trên -> Đồng cấp -> Cấp dưới)."""
    order = {rel: i for i, rel in enumerate(config.RELATIONSHIP_ORDER)}
    ordered = sorted(comments, key=lambda c: order.get(c.get("rel"), 99))
    return "\n".join(
        f"[{config.RELATIONSHIP_DISPLAY.get(c.get('rel'), 'Khác')}] {c['text']}"
        for c in ordered
    )


def _identity_values(structured):
    emp = structured["employee"]
    return {
        "ma_nv":          emp["ma_nv"],
        "ho_ten":         emp["ho_ten"],
        "chuc_danh":      emp["chuc_danh"],
        "ban_chuoi_khoi": emp.get("bo_phan", ""),   # Bộ phận -> Ban/Chuỗi/Khối
        "cap_bac":        emp["cap_bac"],
        "trang_thai":     structured["completion"]["status_text"],
    }


# ---------------------------------------------------------------------------
# Cấu trúc cột (spec) — nguồn duy nhất để dựng header & ghi dữ liệu
# ---------------------------------------------------------------------------
def _build_column_spec():
    """List cột, mỗi cột là dict {kind, ...}. kind ∈ {identity, ketqua, score,
    khuyennghi, ai, review}. Cột 'score' có sub_role ∈ {score, comment}."""
    cols = []
    for key, label in config.IDENTITY_FIELDS_V2:
        cols.append({"kind": "identity", "key": key, "label": label})
    cols.append({"kind": "ketqua", "label": config.LABEL_KET_QUA})

    for block_key, banner, score_src, per_behavior in config.RATER_BLOCKS:
        for idx, (sub_key, behavior_text) in enumerate(config.BEHAVIORS):
            comp_full = config.COMPETENCY_DISPLAY_FULL[sub_key]
            base = {
                "kind": "score", "block_key": block_key, "banner": banner,
                "score_src": score_src, "behavior_index": idx,
                "sub_key": sub_key, "comp_full": comp_full, "behavior_text": behavior_text,
            }
            cols.append({**base, "sub_role": "score"})
            if per_behavior == 2:
                cols.append({**base, "sub_role": "comment"})

    cols.append({"kind": "ykien", "label": config.LABEL_YKIEN})
    cols.append({"kind": "khuyennghi", "label": config.LABEL_KHUYEN_NGHI})
    for key, label, _desc in config.AI_FIELDS:
        cols.append({"kind": "ai", "key": key, "label": label})
    for key, label, _vals in config.REVIEW_FIELDS:
        cols.append({"kind": "review", "key": key, "label": label})
    return cols


_FILL = {
    "identity":   PatternFill("solid", fgColor=config.XLSX_FILL_SOURCE),
    "ketqua":     PatternFill("solid", fgColor=config.XLSX_FILL_SOURCE),
    "score":      PatternFill("solid", fgColor=config.XLSX_FILL_SOURCE),
    "ykien":      PatternFill("solid", fgColor=config.XLSX_FILL_SOURCE),
    "khuyennghi": PatternFill("solid", fgColor=config.XLSX_FILL_AI),
    "ai":         PatternFill("solid", fgColor=config.XLSX_FILL_AI),
    "review":     PatternFill("solid", fgColor=config.XLSX_FILL_REVIEW),
}
_BANNER_FILL = {
    "tong":        PatternFill("solid", fgColor="FF1F4E79"),
    "cap_tren":    PatternFill("solid", fgColor="FF2E5E8C"),
    "dong_nghiep": PatternFill("solid", fgColor="FF3A6EA5"),
    "cap_duoi":    PatternFill("solid", fgColor="FF4A7EB5"),
}


def _merge_runs(cols, key_fn):
    """Sinh (start_idx, end_idx, value) cho các dải cột liên tiếp cùng key_fn
    (bỏ qua cột không phải 'score' -> key None)."""
    runs = []
    i = 0
    n = len(cols)
    while i < n:
        k = key_fn(cols[i])
        if k is None:
            i += 1
            continue
        j = i
        while j + 1 < n and key_fn(cols[j + 1]) == k:
            j += 1
        runs.append((i + 1, j + 1, k[-1] if isinstance(k, tuple) else k))
        i = j + 1
    return runs


def build_competency_workbook(structured_list, out_path):
    """Ghi File thứ 4 (wide, bám đúng mẫu Sheet1). Trả về đường dẫn."""
    cols = _build_column_spec()
    n_cols = len(cols)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = config.BATCH_SHEET_NAME

    white_bold = Font(bold=True, color="FFFFFFFF", size=10)
    bold = Font(bold=True, size=10)
    small = Font(size=8)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # ----- HÀNG 1: banner khối rater (gộp theo block_key liên tiếp) -----
    for start, end, _ in _merge_runs(
        cols, lambda c: ("banner", c["banner"], c["block_key"]) if c["kind"] == "score" else None
    ):
        banner = cols[start - 1]["banner"]
        block_key = cols[start - 1]["block_key"]
        cell = ws.cell(row=1, column=start, value=banner)
        cell.font = white_bold
        cell.alignment = center
        cell.fill = _BANNER_FILL.get(block_key, _BANNER_FILL["tong"])
        if end > start:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)

    # ----- HÀNG 2: tên nhóm năng lực (gộp theo (block_key, sub_key) liên tiếp) -----
    for start, end, _ in _merge_runs(
        cols, lambda c: ("comp", c["block_key"], c["sub_key"]) if c["kind"] == "score" else None
    ):
        cell = ws.cell(row=2, column=start, value=cols[start - 1]["comp_full"])
        cell.font = bold
        cell.alignment = center
        if end > start:
            ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)

    # ----- HÀNG 3: mô tả hành vi (cột điểm) / nhãn (cột khác) -----
    for i, col in enumerate(cols, start=1):
        if col["kind"] == "score":
            if col["sub_role"] == "comment":
                text = config.LABEL_DIEN_GIAI
            else:
                text = col["behavior_text"]
            cell = ws.cell(row=3, column=i, value=text)
            cell.fill = _FILL["score"]
            cell.font = small
            cell.alignment = center
        else:
            cell = ws.cell(row=1, column=i, value=col["label"])
            cell.fill = _FILL[col["kind"]]
            cell.font = bold
            cell.alignment = center
            ws.merge_cells(start_row=1, start_column=i, end_row=3, end_column=i)

    # ----- Bề rộng cột -----
    for i, col in enumerate(cols, start=1):
        letter = get_column_letter(i)
        if col["kind"] == "score":
            ws.column_dimensions[letter].width = 50 if col.get("sub_role") == "comment" else 9
        elif col["kind"] == "identity":
            ws.column_dimensions[letter].width = 20 if col["key"] == "ho_ten" else 15
        elif col["kind"] == "ai":
            ws.column_dimensions[letter].width = 32
        elif col["kind"] == "ykien":
            ws.column_dimensions[letter].width = 48
        else:
            ws.column_dimensions[letter].width = 20

    # ----- DỮ LIỆU từ hàng 4 -----
    review_defaults = {"trang_thai_ra_soat": config.REVIEW_STATUS_DEFAULT}
    first = 4
    for r_off, structured in enumerate(structured_list):
        r = first + r_off
        ident = _identity_values(structured)
        scores24 = _match_behavior_scores(structured)
        ykien_text = "\n".join(f"[{c['rel']}] {c['text']}"
                               for c in structured.get("all_comments", []))
        for i, col in enumerate(cols, start=1):
            kind = col["kind"]
            if kind == "identity":
                value = ident.get(col["key"], "")
            elif kind == "ketqua":
                value = _fmt(structured["total_360"])
            elif kind == "score":
                # Cột điểm -> số; cột "diễn giải" -> ý kiến theo từng mục tiêu (per-objective).
                bscore = scores24[col["behavior_index"]]
                if col["sub_role"] == "score":
                    value = _fmt(bscore[col["score_src"]])
                else:
                    value = _format_behavior_comments(bscore.get("comments", []))
            elif kind == "ykien":
                value = ykien_text
            elif kind == "review":
                value = review_defaults.get(col["key"], "")
            else:  # khuyennghi, ai
                value = ""
            cell = ws.cell(row=r, column=i, value=value)
            cell.fill = _FILL[kind]
            if kind in ("ai", "khuyennghi", "review", "ykien") or (kind == "score" and col.get("sub_role") == "comment"):
                cell.alignment = left_top

    # ----- Dropdown trạng thái rà soát -----
    n_last = first + len(structured_list) - 1
    for i, col in enumerate(cols, start=1):
        if col["kind"] == "review":
            valid = next((v for k, _l, v in config.REVIEW_FIELDS if k == col["key"] and v), None)
            if valid:
                letter = get_column_letter(i)
                dv = DataValidation(type="list", formula1='"' + ",".join(valid) + '"', allow_blank=True)
                dv.add(f"{letter}{first}:{letter}{max(n_last, first)}")
                ws.add_data_validation(dv)

    # ----- Cố định header + cột định danh; auto-filter trên hàng 3 -----
    ws.freeze_panes = ws.cell(row=first, column=len(config.IDENTITY_FIELDS_V2) + 1)
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[3].height = 90
    ws.auto_filter.ref = f"A3:{get_column_letter(n_cols)}{max(n_last, 3)}"

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# Nhập kết quả AI (CSV) -> ghi vào các cột AI của File thứ 4
# ---------------------------------------------------------------------------
def merge_ai_csv(file4_path, csv_path):
    """
    Đọc CSV kết quả AI (do Claude xuất theo prompt chung) và GHI vào đúng các cột
    AI trong File thứ 4, khớp theo 'ma_nv'. Trả về dict thống kê.

    CSV cần có dòng tiêu đề gồm 'ma_nv' + các key cột AI (config.AI_FIELDS).
    """
    spec = _build_column_spec()
    ma_col = next(i for i, c in enumerate(spec)
                  if c["kind"] == "identity" and c["key"] == "ma_nv")
    ai_cols = {c["key"]: i for i, c in enumerate(spec) if c["kind"] == "ai"}

    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        by_ma = {}
        for row in reader:
            ma = (row.get("ma_nv") or "").strip()
            if ma:
                by_ma[ma] = row
        csv_keys = set(reader.fieldnames or [])

    wb = openpyxl.load_workbook(file4_path)
    ws = wb.active
    updated, unknown_cols = 0, sorted(csv_keys - {"ma_nv"} - set(ai_cols))
    for r in range(4, ws.max_row + 1):
        ma_cell = ws.cell(r, ma_col + 1).value
        ma = str(ma_cell).strip() if ma_cell is not None else ""
        src = by_ma.get(ma)
        if not src:
            continue
        for key, ci in ai_cols.items():
            if key in src and src[key] is not None:
                ws.cell(r, ci + 1, value=str(src[key]).strip())
        updated += 1
    wb.save(file4_path)
    return {"rows_in_csv": len(by_ma), "updated": updated, "unknown_columns": unknown_cols}
